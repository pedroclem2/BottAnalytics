"""Read the CSAT workbook into typed :class:`RawRow` records."""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

VALID_SHEETS: tuple[str, ...] = ("3.0 Survey_2025_Data", "4.0 Survey_2026_Data")
INVALID_SHEETS: tuple[str, ...] = (
    "5.0 Invalid survey 2025 Data",
    "6.0 Invalid Survey 2026 Data",
)

_HEADERS: dict[str, str] = {
    "instance": "instance_id",
    "source": "source",
    "tags": "tags",
    "category": "category",
    "metric": "question_text",
    "actual value": "actual_value",
    "created": "created_at",
    "assessment group": "assessment_group",
    "additional information": "additional_information",
    "created by": "agent_email",
    "scaled value": "scaled_value_k",
    "string value": "label_text",
    "assigned to": "agent_name",
    "assignment group": "team_name",
    "parent": "parent",
    "adge": "entity_name",
    "count": "count",
}


@dataclass(frozen=True)
class RawRow:
    """Untyped-but-named projection of one workbook row."""

    instance_id: str | None
    created_at: datetime | None
    entity_name: str | None
    team_name: str | None
    team_parent_l1: str | None
    team_parent_l2: str | None
    agent_name: str | None
    agent_email: str | None
    question_text: str | None
    label_text: str | None
    scaled_value: int | None
    actual_value: int | None
    source_sheet: str


def _coerce_int(value: Any) -> int | None:
    """Best-effort integer coercion for the 1-5 score columns."""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _coerce_str(value: Any) -> str | None:
    """Strip whitespace and return ``None`` for blank values."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _coerce_dt(value: Any) -> datetime | None:
    """Pass-through coercion for datetime cells."""
    if isinstance(value, datetime):
        return value
    return None


def read_sheet(path: Path, sheet_name: str) -> Iterator[RawRow]:
    """Yield :class:`RawRow` records from ``sheet_name`` in the workbook at ``path``.

    The workbook has two columns named ``Parent`` (cols O and Q) and two columns
    named ``Scaled value`` (cols K and R); we keep both by positional ordering.

    Args:
        path: Filesystem path to the ``.xlsm`` workbook.
        sheet_name: Tab name to read.

    Yields:
        One :class:`RawRow` per non-empty data row.
    """
    workbook = load_workbook(filename=path, data_only=True, read_only=True, keep_vba=False)
    try:
        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_indices = _map_headers(header_row)

        # The workbook has duplicate headers; resolve by ordinal occurrence.
        first_parent = header_indices.get("parent_1")
        second_parent = header_indices.get("parent_2")
        first_scaled = header_indices.get("scaled_value_1")
        second_scaled = header_indices.get("scaled_value_2")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None for cell in row):
                continue
            yield RawRow(
                instance_id=_coerce_str(_get(row, header_indices.get("instance_id"))),
                created_at=_coerce_dt(_get(row, header_indices.get("created_at"))),
                entity_name=_coerce_str(_get(row, header_indices.get("entity_name"))),
                team_name=_coerce_str(_get(row, header_indices.get("team_name"))),
                team_parent_l1=_coerce_str(_get(row, first_parent)),
                team_parent_l2=_coerce_str(_get(row, second_parent)),
                agent_name=_coerce_str(_get(row, header_indices.get("agent_name"))),
                agent_email=_coerce_str(_get(row, header_indices.get("agent_email"))),
                question_text=_coerce_str(_get(row, header_indices.get("question_text"))),
                label_text=_coerce_str(_get(row, header_indices.get("label_text"))),
                scaled_value=_coerce_int(_get(row, second_scaled or first_scaled)),
                actual_value=_coerce_int(_get(row, header_indices.get("actual_value"))),
                source_sheet=sheet_name,
            )
    finally:
        workbook.close()


def _get(row: tuple[Any, ...], idx: int | None) -> Any:
    """Safely index into a tuple, returning ``None`` for missing positions."""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _map_headers(header_row: tuple[Any, ...]) -> dict[str, int]:
    """Map workbook headers to column indices, deduplicating ``Parent`` & ``Scaled value``.

    Args:
        header_row: Tuple of header cells from row 1.

    Returns:
        Mapping from logical column key to column index.
    """
    mapping: dict[str, int] = {}
    parent_seen = 0
    scaled_seen = 0
    for idx, raw_header in enumerate(header_row):
        if raw_header is None:
            continue
        header = str(raw_header).strip().lower()
        if header == "parent":
            parent_seen += 1
            mapping[f"parent_{parent_seen}"] = idx
            continue
        if header == "scaled value":
            scaled_seen += 1
            mapping[f"scaled_value_{scaled_seen}"] = idx
            continue
        logical = _HEADERS.get(header)
        if logical is not None:
            mapping[logical] = idx
    return mapping
